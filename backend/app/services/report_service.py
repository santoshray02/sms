"""
Report Generation Service

Generates various reports for school compliance and management:
- RTE Compliance Reports
- CBSE Mandatory Reports
- Financial Reports
- Academic Reports

Supports both PDF and Excel formats.
"""
from datetime import datetime, date
from typing import List, Dict, Any, Optional
from io import BytesIO
from sqlalchemy import select, func, and_, or_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload

# PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.models.student import Student
from app.models.fee import MonthlyFee, FeeStructure
from app.models.payment import Payment
from app.models.academic import AcademicYear, Class, TransportRoute
from app.models.concession import Concession


class ReportService:
    """Service for generating various school reports"""

    @staticmethod
    async def generate_rte_compliance_report(
        db: AsyncSession,
        academic_year_id: int,
        format: str = "pdf"
    ) -> BytesIO:
        """
        Generate RTE Compliance Report

        Includes:
        - EWS/DG student list
        - Category-wise breakdown
        - Concession summary
        """
        # Get academic year
        academic_year = await db.get(AcademicYear, academic_year_id)

        # Get all students with concessions
        stmt = (
            select(Student)
            .where(
                and_(
                    Student.academic_year_id == academic_year_id,
                    Student.status == "active"
                )
            )
            .options(
                joinedload(Student.class_),
                joinedload(Student.concessions)
            )
        )
        result = await db.execute(stmt)
        students = result.unique().scalars().all()

        # Categorize students
        category_counts = {}
        ews_dg_students = []

        for student in students:
            # Count by category
            category = student.category or "General"
            category_counts[category] = category_counts.get(category, 0) + 1

            # EWS/DG students
            if category in ["EWS", "DG"]:
                ews_dg_students.append(student)

        # Get concession summary
        stmt = (
            select(Concession)
            .join(Student)
            .where(Student.academic_year_id == academic_year_id)
        )
        result = await db.execute(stmt)
        concessions = result.scalars().all()

        total_concession_amount = sum(c.amount for c in concessions)

        # Generate report based on format
        if format == "excel":
            return await ReportService._generate_rte_excel(
                academic_year, students, ews_dg_students, category_counts,
                concessions, total_concession_amount
            )
        else:
            return await ReportService._generate_rte_pdf(
                academic_year, students, ews_dg_students, category_counts,
                concessions, total_concession_amount
            )

    @staticmethod
    async def _generate_rte_pdf(
        academic_year, students, ews_dg_students, category_counts,
        concessions, total_concession_amount
    ) -> BytesIO:
        """Generate RTE report as PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        elements.append(Paragraph("RTE Compliance Report", title_style))
        elements.append(Paragraph(f"Academic Year: {academic_year.name}", styles['Normal']))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}", styles['Normal']))
        elements.append(Spacer(1, 20))

        # Summary statistics
        elements.append(Paragraph("<b>Summary</b>", styles['Heading2']))
        summary_data = [
            ["Total Students", len(students)],
            ["EWS/DG Students", len(ews_dg_students)],
            ["Total Concessions", len(concessions)],
            ["Total Concession Amount", f"Rs. {total_concession_amount/100:.2f}"]
        ]
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # Category-wise breakdown
        elements.append(Paragraph("<b>Category-wise Distribution</b>", styles['Heading2']))
        category_data = [["Category", "Count", "Percentage"]]
        for category, count in sorted(category_counts.items()):
            percentage = (count / len(students) * 100) if students else 0
            category_data.append([category, str(count), f"{percentage:.1f}%"])

        category_table = Table(category_data, colWidths=[2*inch, 1.5*inch, 1.5*inch])
        category_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        elements.append(category_table)
        elements.append(Spacer(1, 20))

        # EWS/DG Student List
        if ews_dg_students:
            elements.append(PageBreak())
            elements.append(Paragraph("<b>EWS/DG Student List</b>", styles['Heading2']))

            student_data = [["Admission No", "Name", "Class", "Category", "Concession %"]]
            for student in ews_dg_students:
                student_data.append([
                    student.admission_number,
                    student.full_name,
                    student.class_.name if student.class_ else "N/A",
                    student.category,
                    f"{student.concession_percentage}%"
                ])

            student_table = Table(student_data, colWidths=[1.2*inch, 2*inch, 1*inch, 1*inch, 1*inch])
            student_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            elements.append(student_table)

        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer

    @staticmethod
    async def _generate_rte_excel(
        academic_year, students, ews_dg_students, category_counts,
        concessions, total_concession_amount
    ) -> BytesIO:
        """Generate RTE report as Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "RTE Compliance"

        # Header styling
        header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws['A1'] = "RTE Compliance Report"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Academic Year: {academic_year.name}"
        ws['A3'] = f"Generated: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}"

        # Summary
        ws['A5'] = "Summary"
        ws['A5'].font = Font(bold=True, size=12)
        ws['A6'] = "Total Students"
        ws['B6'] = len(students)
        ws['A7'] = "EWS/DG Students"
        ws['B7'] = len(ews_dg_students)
        ws['A8'] = "Total Concessions"
        ws['B8'] = len(concessions)
        ws['A9'] = "Total Concession Amount"
        ws['B9'] = f"Rs. {total_concession_amount/100:.2f}"

        # Category breakdown
        row = 11
        ws[f'A{row}'] = "Category-wise Distribution"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        ws[f'A{row}'] = "Category"
        ws[f'B{row}'] = "Count"
        ws[f'C{row}'] = "Percentage"
        for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}']]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        row += 1

        for category, count in sorted(category_counts.items()):
            percentage = (count / len(students) * 100) if students else 0
            ws[f'A{row}'] = category
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f"{percentage:.1f}%"
            row += 1

        # EWS/DG Student List
        if ews_dg_students:
            row += 2
            ws[f'A{row}'] = "EWS/DG Student List"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1

            headers = ["Admission No", "Name", "Class", "Category", "Concession %"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
            row += 1

            for student in ews_dg_students:
                ws.cell(row=row, column=1, value=student.admission_number)
                ws.cell(row=row, column=2, value=student.full_name)
                ws.cell(row=row, column=3, value=student.class_.name if student.class_ else "N/A")
                ws.cell(row=row, column=4, value=student.category)
                ws.cell(row=row, column=5, value=f"{student.concession_percentage}%")
                row += 1

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    async def generate_financial_report(
        db: AsyncSession,
        month: int,
        year: int,
        academic_year_id: int,
        format: str = "pdf"
    ) -> BytesIO:
        """
        Generate Financial Report

        Includes:
        - Monthly collection summary
        - Payment mode analysis
        - Outstanding fees by class
        - Defaulter list (7, 15, 30+ days)
        """
        # Get payments for the month
        stmt = (
            select(Payment)
            .where(
                and_(
                    func.extract('month', Payment.payment_date) == month,
                    func.extract('year', Payment.payment_date) == year
                )
            )
            .options(joinedload(Payment.student).joinedload(Student.class_))
        )
        result = await db.execute(stmt)
        payments = result.unique().scalars().all()

        # Calculate statistics
        total_collection = sum(p.amount for p in payments)
        payment_mode_counts = {}
        payment_mode_amounts = {}

        for payment in payments:
            mode = payment.payment_mode
            payment_mode_counts[mode] = payment_mode_counts.get(mode, 0) + 1
            payment_mode_amounts[mode] = payment_mode_amounts.get(mode, 0) + payment.amount

        # Get outstanding fees
        stmt = (
            select(MonthlyFee)
            .where(
                and_(
                    MonthlyFee.academic_year_id == academic_year_id,
                    MonthlyFee.status.in_(["pending", "partial"]),
                    MonthlyFee.amount_pending > 0
                )
            )
            .options(
                joinedload(MonthlyFee.student).joinedload(Student.class_)
            )
        )
        result = await db.execute(stmt)
        outstanding_fees = result.unique().scalars().all()

        total_outstanding = sum(fee.amount_pending for fee in outstanding_fees)

        # Categorize defaulters by days overdue
        today = date.today()
        defaulters_7 = []
        defaulters_15 = []
        defaulters_30 = []

        for fee in outstanding_fees:
            days_overdue = (today - fee.due_date).days
            if days_overdue >= 30:
                defaulters_30.append((fee, days_overdue))
            elif days_overdue >= 15:
                defaulters_15.append((fee, days_overdue))
            elif days_overdue >= 7:
                defaulters_7.append((fee, days_overdue))

        if format == "excel":
            return await ReportService._generate_financial_excel(
                month, year, payments, total_collection,
                payment_mode_counts, payment_mode_amounts,
                outstanding_fees, total_outstanding,
                defaulters_7, defaulters_15, defaulters_30
            )
        else:
            return await ReportService._generate_financial_pdf(
                month, year, payments, total_collection,
                payment_mode_counts, payment_mode_amounts,
                outstanding_fees, total_outstanding,
                defaulters_7, defaulters_15, defaulters_30
            )

    @staticmethod
    async def _generate_financial_pdf(
        month, year, payments, total_collection,
        payment_mode_counts, payment_mode_amounts,
        outstanding_fees, total_outstanding,
        defaulters_7, defaulters_15, defaulters_30
    ) -> BytesIO:
        """Generate financial report as PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        month_names = ["", "January", "February", "March", "April", "May", "June",
                       "July", "August", "September", "October", "November", "December"]
        elements.append(Paragraph(f"Financial Report - {month_names[month]} {year}", title_style))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}", styles['Normal']))
        elements.append(Spacer(1, 20))

        # Collection summary
        elements.append(Paragraph("<b>Collection Summary</b>", styles['Heading2']))
        summary_data = [
            ["Total Payments", len(payments)],
            ["Total Collection", f"Rs. {total_collection/100:.2f}"],
            ["Outstanding Fees", f"Rs. {total_outstanding/100:.2f}"],
            ["Defaulters (7+ days)", len(defaulters_7)],
            ["Defaulters (15+ days)", len(defaulters_15)],
            ["Defaulters (30+ days)", len(defaulters_30)]
        ]
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # Payment mode analysis
        elements.append(Paragraph("<b>Payment Mode Analysis</b>", styles['Heading2']))
        mode_data = [["Payment Mode", "Count", "Amount", "Percentage"]]
        for mode in sorted(payment_mode_counts.keys()):
            count = payment_mode_counts[mode]
            amount = payment_mode_amounts[mode]
            percentage = (amount / total_collection * 100) if total_collection > 0 else 0
            mode_data.append([
                mode.upper(),
                str(count),
                f"Rs. {amount/100:.2f}",
                f"{percentage:.1f}%"
            ])

        mode_table = Table(mode_data, colWidths=[1.5*inch, 1*inch, 1.5*inch, 1*inch])
        mode_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        elements.append(mode_table)
        elements.append(Spacer(1, 20))

        # Defaulters list (30+ days)
        if defaulters_30:
            elements.append(PageBreak())
            elements.append(Paragraph("<b>Critical Defaulters (30+ days overdue)</b>", styles['Heading2']))

            defaulter_data = [["Student", "Class", "Amount", "Due Date", "Days Overdue"]]
            for fee, days in sorted(defaulters_30, key=lambda x: x[1], reverse=True):
                defaulter_data.append([
                    fee.student.full_name,
                    fee.student.class_.name if fee.student.class_ else "N/A",
                    f"Rs. {fee.amount_pending/100:.2f}",
                    fee.due_date.strftime('%d-%b-%Y'),
                    str(days)
                ])

            defaulter_table = Table(defaulter_data, colWidths=[2*inch, 1*inch, 1.2*inch, 1.2*inch, 1*inch])
            defaulter_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.red),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            elements.append(defaulter_table)

        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer

    @staticmethod
    async def _generate_financial_excel(
        month, year, payments, total_collection,
        payment_mode_counts, payment_mode_amounts,
        outstanding_fees, total_outstanding,
        defaulters_7, defaulters_15, defaulters_30
    ) -> BytesIO:
        """Generate financial report as Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Financial Report"

        # Styling
        header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # Title
        month_names = ["", "January", "February", "March", "April", "May", "June",
                       "July", "August", "September", "October", "November", "December"]
        ws['A1'] = f"Financial Report - {month_names[month]} {year}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Generated: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}"

        # Collection summary
        ws['A4'] = "Collection Summary"
        ws['A4'].font = Font(bold=True, size=12)
        ws['A5'] = "Total Payments"
        ws['B5'] = len(payments)
        ws['A6'] = "Total Collection"
        ws['B6'] = f"Rs. {total_collection/100:.2f}"
        ws['A7'] = "Outstanding Fees"
        ws['B7'] = f"Rs. {total_outstanding/100:.2f}"
        ws['A8'] = "Defaulters (7+ days)"
        ws['B8'] = len(defaulters_7)
        ws['A9'] = "Defaulters (15+ days)"
        ws['B9'] = len(defaulters_15)
        ws['A10'] = "Defaulters (30+ days)"
        ws['B10'] = len(defaulters_30)

        # Payment mode analysis
        row = 12
        ws[f'A{row}'] = "Payment Mode Analysis"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        for col, header in enumerate(["Payment Mode", "Count", "Amount", "Percentage"], start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        row += 1

        for mode in sorted(payment_mode_counts.keys()):
            ws.cell(row=row, column=1, value=mode.upper())
            ws.cell(row=row, column=2, value=payment_mode_counts[mode])
            ws.cell(row=row, column=3, value=f"Rs. {payment_mode_amounts[mode]/100:.2f}")
            percentage = (payment_mode_amounts[mode] / total_collection * 100) if total_collection > 0 else 0
            ws.cell(row=row, column=4, value=f"{percentage:.1f}%")
            row += 1

        # Defaulters list
        if defaulters_30:
            row += 2
            ws[f'A{row}'] = "Critical Defaulters (30+ days overdue)"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1

            for col, header in enumerate(["Student", "Class", "Amount", "Due Date", "Days Overdue"], start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            row += 1

            for fee, days in sorted(defaulters_30, key=lambda x: x[1], reverse=True):
                ws.cell(row=row, column=1, value=fee.student.full_name)
                ws.cell(row=row, column=2, value=fee.student.class_.name if fee.student.class_ else "N/A")
                ws.cell(row=row, column=3, value=f"Rs. {fee.amount_pending/100:.2f}")
                ws.cell(row=row, column=4, value=fee.due_date.strftime('%d-%b-%Y'))
                ws.cell(row=row, column=5, value=days)
                row += 1

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    async def generate_academic_report(
        db: AsyncSession,
        academic_year_id: int,
        format: str = "pdf"
    ) -> BytesIO:
        """
        Generate Academic Report

        Includes:
        - Class-wise strength
        - Gender distribution
        - Category-wise enrollment
        - Transport utilization
        """
        # Get all active students
        stmt = (
            select(Student)
            .where(
                and_(
                    Student.academic_year_id == academic_year_id,
                    Student.status == "active"
                )
            )
            .options(
                joinedload(Student.class_),
                joinedload(Student.transport_route)
            )
        )
        result = await db.execute(stmt)
        students = result.unique().scalars().all()

        # Class-wise statistics
        class_stats = {}
        gender_stats = {"Male": 0, "Female": 0, "Other": 0}
        category_stats = {}
        transport_stats = {"Yes": 0, "No": 0}

        for student in students:
            # Class-wise
            class_name = student.class_.name if student.class_ else "Unassigned"
            if class_name not in class_stats:
                class_stats[class_name] = {"total": 0, "male": 0, "female": 0}
            class_stats[class_name]["total"] += 1
            if student.gender == "Male":
                class_stats[class_name]["male"] += 1
            elif student.gender == "Female":
                class_stats[class_name]["female"] += 1

            # Gender
            gender_stats[student.gender] = gender_stats.get(student.gender, 0) + 1

            # Category
            category = student.category or "General"
            category_stats[category] = category_stats.get(category, 0) + 1

            # Transport
            if student.transport_route_id:
                transport_stats["Yes"] += 1
            else:
                transport_stats["No"] += 1

        if format == "excel":
            return await ReportService._generate_academic_excel(
                students, class_stats, gender_stats, category_stats, transport_stats
            )
        else:
            return await ReportService._generate_academic_pdf(
                students, class_stats, gender_stats, category_stats, transport_stats
            )

    @staticmethod
    async def _generate_academic_pdf(
        students, class_stats, gender_stats, category_stats, transport_stats
    ) -> BytesIO:
        """Generate academic report as PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        elements.append(Paragraph("Academic Enrollment Report", title_style))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}", styles['Normal']))
        elements.append(Spacer(1, 20))

        # Overall summary
        elements.append(Paragraph("<b>Overall Summary</b>", styles['Heading2']))
        summary_data = [
            ["Total Students", len(students)],
            ["Boys", gender_stats.get("Male", 0)],
            ["Girls", gender_stats.get("Female", 0)],
            ["Using Transport", transport_stats["Yes"]]
        ]
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # Class-wise breakdown
        elements.append(Paragraph("<b>Class-wise Strength</b>", styles['Heading2']))
        class_data = [["Class", "Total", "Boys", "Girls", "Boy %", "Girl %"]]
        for class_name in sorted(class_stats.keys()):
            stats = class_stats[class_name]
            boy_pct = (stats["male"] / stats["total"] * 100) if stats["total"] > 0 else 0
            girl_pct = (stats["female"] / stats["total"] * 100) if stats["total"] > 0 else 0
            class_data.append([
                class_name,
                str(stats["total"]),
                str(stats["male"]),
                str(stats["female"]),
                f"{boy_pct:.1f}%",
                f"{girl_pct:.1f}%"
            ])

        class_table = Table(class_data, colWidths=[1.2*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch])
        class_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        elements.append(class_table)
        elements.append(Spacer(1, 20))

        # Category distribution
        elements.append(Paragraph("<b>Category-wise Distribution</b>", styles['Heading2']))
        category_data = [["Category", "Count", "Percentage"]]
        for category in sorted(category_stats.keys()):
            count = category_stats[category]
            percentage = (count / len(students) * 100) if students else 0
            category_data.append([category, str(count), f"{percentage:.1f}%"])

        category_table = Table(category_data, colWidths=[2*inch, 1.5*inch, 1.5*inch])
        category_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        elements.append(category_table)

        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer

    @staticmethod
    async def _generate_academic_excel(
        students, class_stats, gender_stats, category_stats, transport_stats
    ) -> BytesIO:
        """Generate academic report as Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Academic Report"

        # Styling
        header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # Title
        ws['A1'] = "Academic Enrollment Report"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Generated: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}"

        # Overall summary
        ws['A4'] = "Overall Summary"
        ws['A4'].font = Font(bold=True, size=12)
        ws['A5'] = "Total Students"
        ws['B5'] = len(students)
        ws['A6'] = "Boys"
        ws['B6'] = gender_stats.get("Male", 0)
        ws['A7'] = "Girls"
        ws['B7'] = gender_stats.get("Female", 0)
        ws['A8'] = "Using Transport"
        ws['B8'] = transport_stats["Yes"]

        # Class-wise breakdown
        row = 10
        ws[f'A{row}'] = "Class-wise Strength"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        for col, header in enumerate(["Class", "Total", "Boys", "Girls", "Boy %", "Girl %"], start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        row += 1

        for class_name in sorted(class_stats.keys()):
            stats = class_stats[class_name]
            boy_pct = (stats["male"] / stats["total"] * 100) if stats["total"] > 0 else 0
            girl_pct = (stats["female"] / stats["total"] * 100) if stats["total"] > 0 else 0
            ws.cell(row=row, column=1, value=class_name)
            ws.cell(row=row, column=2, value=stats["total"])
            ws.cell(row=row, column=3, value=stats["male"])
            ws.cell(row=row, column=4, value=stats["female"])
            ws.cell(row=row, column=5, value=f"{boy_pct:.1f}%")
            ws.cell(row=row, column=6, value=f"{girl_pct:.1f}%")
            row += 1

        # Category distribution
        row += 2
        ws[f'A{row}'] = "Category-wise Distribution"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        for col, header in enumerate(["Category", "Count", "Percentage"], start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        row += 1

        for category in sorted(category_stats.keys()):
            count = category_stats[category]
            percentage = (count / len(students) * 100) if students else 0
            ws.cell(row=row, column=1, value=category)
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f"{percentage:.1f}%")
            row += 1

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
